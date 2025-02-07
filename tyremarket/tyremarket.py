# import time
# import random
# import re
# import pandas as pd
# from datetime import datetime
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from openpyxl import Workbook, load_workbook
# from webdriver_manager.chrome import ChromeDriverManager

# def scrape_ceat_tyres(url):
#     # Configure Selenium WebDriver
#     chrome_options = Options()
#     chrome_options.add_argument("--headless")  # Run browser in headless mode
#     driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

#     try:
#         # Open the website
#         driver.get(url)

#         # Wait for the page to load and ensure tyre containers are present
#         WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "row")))

#         # Extract tyre containers (parent elements containing all data for a tyre)
#         tyre_containers = driver.find_elements(By.CLASS_NAME, "row")

#         # Extract tyre model names, specifications, and prices
#         tyres = []  # List to store tyre details as dictionaries

#         for container in tyre_containers:
#             try:
#                 # Extract model name
#                 model_element = container.find_element(By.CLASS_NAME, "tmpronameone").text.strip()
#                 model_name = model_element.text.strip()

#                 # Extract specification
#                 spec_element = container.find_element(By.CLASS_NAME, "tmpronametwo").text.strip()
#                 specification = spec_element.text.strip()

#                 # Extract price
#                 price_element = container.find_element(By.CLASS_NAME, "ProductListprice")
#                 price = price_element.text.strip()

#                 # Extract link
#                 link = model_element.get_attribute("href")

#                 braking_element = container.find_element(By.CLASS_NAME, "prodtlblname1")  # Braking
#                 grip_element = container.find_element(By.CLASS_NAME, "prodtlblname2")    # Excellent dry and wet grip
#                 scalability_element = container.find_element(By.CLASS_NAME, "prodtlblname3") # Scalability

#                 # Create a list of all three elements (Braking, Grip, Scalability, Long Life)
#                 elements = [braking_element.text.strip() if braking_element else None,
#                             grip_element.text.strip() if grip_element else None,
#                             scalability_element.text.strip() if scalability_element else None]

#                 # Check for keywords and store in respective columns
#                 braking = None
#                 grip = None
#                 scalable = None
#                 long_life = None
#                 ride = None

#                 # Iterate over the elements and check for keywords
#                 for elem in elements:
#                     if elem:
#                         if "Braking" in elem:
#                             braking = elem
#                         elif "Excellent Dry & Wet Grip" in elem:
#                             grip = elem
#                         elif "Stable" in elem:
#                             scalable = elem
#                         elif "Long life" in elem:
#                             long_life = elem
#                         elif "Ride" in elem:
#                             ride = elem


#                 # Determine if "A/T", "L/T", "LT", or "UHP" is in the model name
#                 at_type = "N/A"
#                 lt_type = "N/A"
#                 uhp_type = "N/A"

#                 if "A/T" in model_name or "AT" in model_name:
#                     at_type = "A/T"
#                 if "L/T" in model_name or "LT" in model_name:
#                     lt_type = "L/T"
#                 if "UHP" in model_name:
#                     uhp_type = "UHP"

#                 # Append the data to the list if all fields are valid
#                 if model_name and specification and price and link:
#                     tyres.append({
#                         "model_name": model_name,
#                         "specification": specification,
#                         "link": link,
#                         "price": price,
#                         "braking": braking,
#                         "grip": grip,
#                         "scalable": scalable,
#                         "long_life": long_life,
#                         "ride": ride,
#                         "at_type": at_type,
#                         "lt_type": lt_type,
#                         "uhp_type": uhp_type,
#                     })
#             except Exception as e:
#                 print(f"Error extracting data for a container: {e}")

#         # Save the tyre data to an Excel file
#         save_models_to_excel(tyres)

#     except Exception as e:
#         print(f"An error occurred while scraping the CEAT brand page: {e}")

#     finally:
#         driver.quit()

# def save_models_to_excel(tyres):
#     file_name = "tyremarket.xlsx"

#     try:
#         # Load the workbook or create a new one if it doesn't exist
#         try:
#             workbook = load_workbook(file_name)
#         except FileNotFoundError:
#             print(f"File '{file_name}' not found. Creating a new one.")
#             workbook = Workbook()

#         sheet = workbook.active
#         sheet.title = "Tyre Data"

#         # Set headers if the file is newly created or headers are missing
#         if sheet.max_row == 1:
#             sheet["A1"] = "Model Name"
#             sheet["B1"] = "Specification"
#             sheet["C1"] = "Profile"
#             sheet["D1"] = "Width"
#             sheet["F1"] = "Contaminant Radius"
#             sheet["G1"] = "Load Index"
#             sheet["H1"] = "Speed Rating"
#             sheet["I1"] = "Type"
#             sheet["AF1"] = "Price"
#             sheet["AG1"] = "URL"
#             sheet["V1"] = "Ride"
#             sheet["W1"] = "Braking"
#             sheet["X1"] = "Excellent Grip"
#             sheet["Y1"] = "Scalable"
#             sheet["Z1"] = "Long Life"
#             sheet["S1"] = "A/T"
#             sheet["T1"] = "L/T"
#             sheet["U1"] = "UHP"

#         # Write each tyre data row into respective columns
#         for row, tyre in enumerate(tyres, start=2):
#             model = tyre.get("model_name")
#             specification = tyre.get("specification")
#             price = tyre.get("price")
#             link = tyre.get("link")

#             braking = tyre.get("braking")
#             grip = tyre.get("grip")
#             scalable = tyre.get("scalable")
#             long_life = tyre.get("long_life")
#             ride = tyre.get("ride")

#             at_type = tyre.get("at_type")
#             lt_type = tyre.get("lt_type")
#             uhp_type = tyre.get("uhp_type")

#             # Default values for parsed details
#             profile, width, radius, tyre_type, load_index, speed_rating = None, None, None, None, None, None

#             # Parse the specification string
#             if specification:
#                 try:
#                     # Extract profile and width (e.g., "185/70")
#                     match = re.search(r"(\d+)/(\d+)", specification)
#                     if match:
#                         profile = int(match.group(1))  # Convert profile to integer
#                         width = int(match.group(2))    # Convert width to integer

#                     # Extract radius (e.g., "R 14")
#                     match = re.search(r"R\s*(\d+)", specification)
#                     if match:
#                         radius = int(match.group(1))  # Convert radius to integer

#                     # Identify tyre type (Tubeless or Tube-type)
#                     if "TUBELESS" in specification.upper():
#                         tyre_type = "Tubeless"
#                     elif "TUBE-TYPE" in specification.upper():
#                         tyre_type = "Tube-type"
#                     elif "REQUIRES TUBE" in specification.upper():
#                         tyre_type = "Requires Tube"

#                     # Extract load index and speed rating (e.g., "88 T" or "88T") after tyre type
#                     match = re.search(r"(TUBELESS|TUBE-TYPE|Requires Tube).*?(\d+)\s?([A-Z])", specification, re.IGNORECASE)
#                     if match:
#                         load_index = int(match.group(2))  # Convert load index to integer
#                         speed_rating = match.group(3)  # Speed rating (e.g., T)
#                 except Exception as e:
#                     print(f"Error parsing specification for '{specification}': {e}")

#             # Save details into respective columns
#             sheet[f"A{row}"] = model
#             sheet[f"B{row}"] = specification
#             sheet[f"C{row}"] = profile
#             sheet[f"D{row}"] = width
#             sheet[f"F{row}"] = radius
#             sheet[f"G{row}"] = load_index
#             sheet[f"H{row}"] = speed_rating
#             sheet[f"I{row}"] = tyre_type

#             # Price conversion: Clean the price string and convert to integer
#             try:
#                 clean_price = re.sub(r'[^\d]', '', price)  # Remove non-numeric characters
#                 price_value = int(clean_price) if clean_price.isdigit() else None
#                 sheet[f"AF{row}"] = price_value if price_value is not None else "N/A"
#             except Exception as e:
#                 print(f"Error converting price '{price}': {e}")
#                 sheet[f"AF{row}"] = "N/A"

#             sheet[f"AG{row}"] = link

#             # Add the new data to the respective columns
#             sheet[f"V{row}"] = ride if ride else "N/A"
#             sheet[f"W{row}"] = braking if braking else "N/A"
#             sheet[f"X{row}"] = grip if grip else "N/A"
#             sheet[f"Y{row}"] = scalable if scalable else "N/A"
#             sheet[f"Z{row}"] = long_life if long_life else "N/A"
#             sheet[f"S{row}"] = at_type
#             sheet[f"T{row}"] = lt_type
#             sheet[f"U{row}"] = uhp_type

#         # Save the workbook
#         workbook.save(file_name)
#         print(f"Data successfully saved to {file_name}")

#     except Exception as e:
#         print(f"Error saving tyre data to Excel: {e}")

# if __name__ == "__main__":

#     url=input("enter url:--> ")
#     scrape_ceat_tyres(url)






import time
import random
import re
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from webdriver_manager.chrome import ChromeDriverManager

# Function to read links from an Excel file
def read_links_from_excel(file_name="input_link.xlsx"):
    try:
        df = pd.read_excel(file_name)  # Read Excel file
        return df['Links'].dropna().tolist()  # Extract the column named 'Links' and convert to list
    except Exception as e:
        print(f"Error reading links from Excel: {e}")
        return []

# Function to scrape tyre data from a given URL
def scrape_ceat_tyres(url):
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Run browser in headless mode
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    tyres = []

    try:
        driver.get(url)

        # Wait for the page to load
        WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "row")))
        tyre_containers = driver.find_elements(By.CLASS_NAME, "row")

        for container in tyre_containers:
            try:
                model_name = container.find_element(By.CLASS_NAME, "tmpronameone").text.strip()
                specification = container.find_element(By.CLASS_NAME, "tmpronametwo").text.strip()
                price = container.find_element(By.CLASS_NAME, "ProductListprice").text.strip()
                link = container.find_element(By.CLASS_NAME, "tmpronameone").get_attribute("href")


                braking_element = container.find_element(By.CLASS_NAME, "prodtlblname1")  # Braking
                grip_element = container.find_element(By.CLASS_NAME, "prodtlblname2")    # Excellent dry and wet grip
                scalability_element = container.find_element(By.CLASS_NAME, "prodtlblname3") # Scalabilit
            #     Create a list of all three elements (Braking, Grip, Scalability, Long Life)
                elements = [braking_element.text.strip() if braking_element else None,
                            grip_element.text.strip() if grip_element else None,
                            scalability_element.text.strip() if scalability_element else None
                ]
            #     Check for keywords and store in respective columns
                braking = None
                grip = None
                scalable = None
                long_life = None
                ride = None
            #     Iterate over the elements and check for keywords
                for elem in elements:
                    elem=elem.lower()
                    if elem:
                        if "braking" in elem:
                            braking = elem
                        elif "excellent dry & wet grip" in elem:
                            grip = elem
                        elif "stable" in elem:
                            scalable = elem
                        elif "life" in elem:
                            long_life = elem
                        elif "ride" in elem:
                            ride = elem

                # Type detection
                at_type, lt_type, uhp_type = "N/A", "N/A", "N/A"
                if "A/T" in model_name or "AT" in model_name:
                    at_type = "A/T"
                if "L/T" in model_name or "LT" in model_name:
                    lt_type = "L/T"
                if "UHP" in model_name:
                    uhp_type = "UHP"

                tyres.append({
                    "model_name": model_name,
                    "specification": specification,
                    "link": link,
                    "price": price,
                    "braking": braking,
                    "grip": grip,
                    "scalable": scalable,
                    "long_life": long_life,
                    "ride": ride,
                    "at_type": at_type,
                    "lt_type": lt_type,
                    "uhp_type": uhp_type,
                })

            except Exception as e:
                print(f"Error extracting data: {e}")

    except Exception as e:
        print(f"Error loading page: {e}")

    finally:
        driver.quit()

    return tyres

# Function to save tyre data to an Excel file
def save_models_to_excel(tyres, file_name="tyremarket.xlsx"):
    try:
        # Load the workbook or create a new one if it doesn't exist
        try:
            workbook = load_workbook(file_name)
        except FileNotFoundError:
            print(f"File '{file_name}' not found. Creating a new one.")
            workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Tyre Data"

        # Set headers if the file is newly created or headers are missing
        if sheet.max_row == 1:
            sheet["A1"] = "Model Name"
            sheet["B1"] = "Specification"
            sheet["C1"] = "Profile"
            sheet["D1"] = "Width"
            sheet["F1"] = "Contaminant Radius"
            sheet["G1"] = "Load Index"
            sheet["H1"] = "Speed Rating"
            sheet["I1"] = "Type"
            sheet["AF1"] = "Price"
            sheet["AG1"] = "URL"
            sheet["V1"] = "Ride"
            sheet["W1"] = "Braking"
            sheet["X1"] = "Excellent Grip"
            sheet["Y1"] = "Scalable"
            sheet["Z1"] = "Long Life"
            sheet["S1"] = "A/T"
            sheet["T1"] = "L/T"
            sheet["U1"] = "UHP"

        # Write each tyre data row into respective columns
        for row, tyre in enumerate(tyres, start=2):
            model = tyre.get("model_name")
            specification = tyre.get("specification")
            price = tyre.get("price")
            link = tyre.get("link")

            braking = tyre.get("braking")
            grip = tyre.get("grip")
            scalable = tyre.get("scalable")
            long_life = tyre.get("long_life")
            ride = tyre.get("ride")

            at_type = tyre.get("at_type")
            lt_type = tyre.get("lt_type")
            uhp_type = tyre.get("uhp_type")

            # Default values for parsed details
            profile, width, radius, tyre_type, load_index, speed_rating = None, None, None, None, None, None

            # Parse the specification string
            if specification:
                try:
                    # Extract profile and width (e.g., "185/70")
                    match = re.search(r"(\d+)/(\d+)", specification)
                    if match:
                        profile = int(match.group(1))  # Convert profile to integer
                        width = int(match.group(2))    # Convert width to integer

                    # Extract radius (e.g., "R 14")
                    match = re.search(r"R\s*(\d+)", specification)
                    if match:
                        radius = int(match.group(1))  # Convert radius to integer

                    # Identify tyre type (Tubeless or Tube-type)
                    if "TUBELESS" in specification.upper():
                        tyre_type = "Tubeless"
                    elif "TUBE-TYPE" in specification.upper():
                        tyre_type = "Tube-type"
                    elif "REQUIRES TUBE" in specification.upper():
                        tyre_type = "Requires Tube"

                    # Extract load index and speed rating (e.g., "88 T" or "88T") after tyre type
                    match = re.search(r"(TUBELESS|TUBE-TYPE|Requires Tube).*?(\d+)\s?([A-Z])", specification, re.IGNORECASE)
                    if match:
                        load_index = int(match.group(2))  # Convert load index to integer
                        speed_rating = match.group(3)  # Speed rating (e.g., T)
                except Exception as e:
                    print(f"Error parsing specification for '{specification}': {e}")

            # Save details into respective columns
            sheet[f"A{row}"] = model
            sheet[f"B{row}"] = specification
            sheet[f"C{row}"] = profile
            sheet[f"D{row}"] = width
            sheet[f"F{row}"] = radius
            sheet[f"G{row}"] = load_index
            sheet[f"H{row}"] = speed_rating
            sheet[f"I{row}"] = tyre_type

            # Price conversion: Clean the price string and convert to integer
            try:
                clean_price = re.sub(r'[^\d]', '', price)  # Remove non-numeric characters
                price_value = int(clean_price) if clean_price.isdigit() else None
                sheet[f"AF{row}"] = price_value if price_value is not None else "N/A"
            except Exception as e:
                print(f"Error converting price '{price}': {e}")
                sheet[f"AF{row}"] = "N/A"

            sheet[f"AG{row}"] = link

            # Add the new data to the respective columns
            sheet[f"V{row}"] = ride if ride else "N/A"
            sheet[f"W{row}"] = braking if braking else "N/A"
            sheet[f"X{row}"] = grip if grip else "N/A"
            sheet[f"Y{row}"] = scalable if scalable else "N/A"
            sheet[f"Z{row}"] = long_life if long_life else "N/A"
            sheet[f"S{row}"] = at_type
            sheet[f"T{row}"] = lt_type
            sheet[f"U{row}"] = uhp_type

        # Save the workbook
        workbook.save(file_name)
        print(f"Data successfully saved to {file_name}")

    except Exception as e:
        print(f"Error saving tyre data to Excel: {e}")

if __name__ == "__main__":
    input_file = "input_links.xlsx"
    urls = read_links_from_excel(input_file)
    count=1

    if not urls:
        print("No valid links found in the input file.")
    else:
        all_tyres = []

        for url in urls:
            print(f"{count}. Scraping: {url}")
            count=count+1
            time.sleep(random.randint(3,6))  # Random delay to avoid blocking
            tyres = scrape_ceat_tyres(url)
            all_tyres.extend(tyres)

        if all_tyres:
            save_models_to_excel(all_tyres)
        else:
            print("No data scraped.")
